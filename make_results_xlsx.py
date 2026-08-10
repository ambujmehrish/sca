"""Build a POLISHED HyperAlign_results.xlsx — zero-shot + finetuning, Ours/GRAM-base/paper, R@1/R@10.
Styled: title banners, grouped headers, colored deltas, zebra rows, borders, frozen headers."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook(); wb.remove(wb.active)

# palette
INDIGO="3F37C9"; INK="1F2430"; GREYH="E9ECF4"; ZEBRA="F6F7FB"
GRN="1E8A5B"; GRNBG="DFF3E8"; RED="C23A4B"; REDBG="FBE1E5"; TIE="7A8496"; TIEBG="EEF1F6"
FN="Times New Roman"                      # font family
thin=Side(style="thin",color="D5DAE6")
BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
CEN=Alignment(horizontal="center",vertical="center")
LEFT=Alignment(horizontal="left",vertical="center")
TITLE_F=Font(name=FN,bold=True,size=16,color="FFFFFF")
HDR_F=Font(name=FN,bold=True,size=13,color="FFFFFF")
OURS_F=Font(name=FN,bold=True,size=13,color=INK)
NUM_F=Font(name=FN,size=13,color=INK)
TITLE_FILL=PatternFill("solid",fgColor=INDIGO)
HDR_FILL=PatternFill("solid",fgColor="5B54C9")
ZEBRA_FILL=PatternFill("solid",fgColor=ZEBRA)

def dfill(v):
    try: v=float(v)
    except: return None,None
    if v>0.2: return Font(name=FN,bold=True,size=13,color=GRN),PatternFill("solid",fgColor=GRNBG)
    if v<-0.2: return Font(name=FN,bold=True,size=13,color=RED),PatternFill("solid",fgColor=REDBG)
    return Font(name=FN,bold=True,size=13,color=TIE),PatternFill("solid",fgColor=TIEBG)

def build(name,title,groups,subhdr,rows,ours_cols,delta_cols,merge_first=None):
    ws=wb.create_sheet(name)
    ncol=len(subhdr)
    # row1 title banner
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncol)
    t=ws.cell(1,1,title); t.font=TITLE_F; t.fill=TITLE_FILL; t.alignment=LEFT
    ws.row_dimensions[1].height=24
    # row2 group headers (merged)
    c=1
    for label,span in groups:
        if label:
            ws.merge_cells(start_row=2,start_column=c,end_row=2,end_column=c+span-1)
            g=ws.cell(2,c,label); g.font=HDR_F; g.fill=HDR_FILL; g.alignment=CEN; g.border=BORD
        else:
            g=ws.cell(2,c,""); g.fill=HDR_FILL
        for cc in range(c,c+span): ws.cell(2,cc).fill=HDR_FILL; ws.cell(2,cc).border=BORD
        c+=span
    # row3 sub headers
    for j,h in enumerate(subhdr,1):
        cell=ws.cell(3,j,h); cell.font=HDR_F; cell.fill=HDR_FILL; cell.alignment=CEN; cell.border=BORD
    # data
    r0=4
    for i,row in enumerate(rows):
        zebra = ZEBRA_FILL if i%2 else None
        for j,v in enumerate(row,1):
            cell=ws.cell(r0+i,j,v); cell.border=BORD
            cell.alignment=LEFT if j<=2 else CEN
            if zebra: cell.fill=zebra
            if j in ours_cols: cell.font=OURS_F
            elif j>2: cell.font=NUM_F
            if j==1: cell.font=Font(name=FN,bold=True,size=13,color=INK)
            if j in delta_cols:
                f,fl=dfill(v)
                if f: cell.font=f
                if fl: cell.fill=fl
    # merge first col by benchmark (visual grouping)
    if merge_first:
        start=r0; prev=rows[0][0]
        for i in range(1,len(rows)+1):
            cur=rows[i][0] if i<len(rows) else None
            if cur!=prev:
                if r0+i-1>start: ws.merge_cells(start_row=start,start_column=1,end_row=r0+i-1,end_column=1)
                ws.cell(start,1).alignment=Alignment(horizontal="left",vertical="center")
                start=r0+i; prev=cur
    ws.freeze_panes="A4"
    # row heights (readable with bigger font)
    ws.row_dimensions[1].height=28; ws.row_dimensions[2].height=20; ws.row_dimensions[3].height=20
    for i in range(len(rows)): ws.row_dimensions[r0+i].height=21
    # widths
    for col in range(1,ncol+1):
        L=get_column_letter(col)
        mx=max([len(str(ws.cell(rr,col).value or "")) for rr in range(3,r0+len(rows))]+[6])
        ws.column_dimensions[L].width=mx+3 if col<=2 else max(mx+2,11)
    ws.sheet_view.showGridLines=False
    return ws

# ============ Zero-shot T2V ============
zs_t2v=[
 ["MSR-VTT","T-V",52.3,82.1,51.8,82.8,52.8,82.9,0.5],
 ["MSR-VTT","T-VA",53.2,82.9,52.8,83.0,54.2,83.9,0.4],
 ["MSR-VTT","T-VAS",53.2,83.3,53.4,83.3,54.8,82.9,-0.2],
 ["DiDeMo","T-V",50.4,76.4,51.1,77.6,54.0,80.7,-0.7],
 ["DiDeMo","T-VA",50.3,76.7,50.9,77.6,54.2,79.3,-0.6],
 ["ActivityNet","T-V",58.6,89.1,60.3,90.2,58.9,91.2,-1.7],
 ["ActivityNet","T-VA",59.4,89.1,62.0,90.6,59.0,91.1,-2.6],
 ["VATEX","T-V",78.7,97.7,79.6,97.9,81.1,99.5,-0.9],
 ["VATEX","T-VA",80.6,98.4,81.4,98.4,83.9,98.6,-0.8],
 ["VATEX","T-VAS",79.9,98.5,79.7,98.5,83.5,98.8,0.2],
]
build("Zero-shot T2V","Zero-shot  ·  Text → Video retrieval  (R@1 / R@10)",
 [("",2),("Ours",2),("GRAM-base",2),("Paper",2),("",1)],
 ["Benchmark","Mode","R@1","R@10","R@1","R@10","R@1","R@10","Δ HG"],
 zs_t2v, ours_cols={3,4}, delta_cols={9}, merge_first=True)

# ============ Zero-shot V2T ============
zs_v2t=[
 ["MSR-VTT","T-V",49.0,79.8,48.9,82.2,49.5,81.7],
 ["MSR-VTT","T-VA",49.5,80.5,49.2,82.4,50.5,82.2],
 ["MSR-VTT","T-VAS",52.4,80.9,51.9,82.0,52.9,82.9],
 ["DiDeMo","T-V",48.1,75.7,49.2,77.1,52.3,80.3],
 ["DiDeMo","T-VA",48.8,76.3,49.0,76.9,52.2,78.9],
 ["ActivityNet","T-V",52.4,85.9,53.9,87.0,50.9,85.4],
 ["ActivityNet","T-VA",52.4,85.6,54.2,87.1,50.4,85.8],
 ["VATEX","T-V",77.1,96.0,76.1,96.0,79.0,98.3],
 ["VATEX","T-VA",79.7,97.2,77.8,96.5,79.2,99.0],
 ["VATEX","T-VAS",78.1,96.9,78.3,96.3,82.7,98.1],
]
build("Zero-shot V2T","Zero-shot  ·  Video → Text retrieval  (R@1 / R@10)",
 [("",2),("Ours",2),("GRAM-base",2),("Paper",2)],
 ["Benchmark","Mode","R@1","R@10","R@1","R@10","R@1","R@10"],
 zs_v2t, ours_cols={3,4}, delta_cols=set(), merge_first=True)

# ============ Zero-shot Audio ============
aud=[
 ["AudioCaps (T-V-A)","R@1 / R@10",33.6,72.7,33.2,74.1,33.2,75.3,0.4],
 ["VGGSound 5K (T-A-V)","Acc@1 / Acc@10",37.0,74.0,40.8,77.7,40.6,78.1,-3.8],
]
build("Zero-shot Audio","Zero-shot  ·  Audio  (AudioCaps R@ · VGGSound Acc)",
 [("",2),("Ours",2),("GRAM-base",2),("Paper",2),("",1)],
 ["Benchmark","Metric","@1","@10","@1","@10","@1","@10","Δ HG"],
 aud, ours_cols={3,4}, delta_cols={9})

# ============ Finetuning ============
ft=[
 ["MSR-VTT","T-V",56.3,87.0,55.7,86.4,56.4,88.1,56.4,87.6,0.6],
 ["MSR-VTT","T-VA",60.9,89.4,58.4,87.0,58.9,91.2,59.0,89.1,2.5],
 ["MSR-VTT","T-VAS",62.6,89.7,64.0,89.3,61.2,92.2,64.8,91.5,-1.4],
 ["DiDeMo","T-V",62.1,85.6,66.4,89.9,59.1,87.4,63.2,91.6,-4.3],
 ["DiDeMo","T-VA",61.4,86.5,67.3,90.1,60.4,88.8,63.5,91.4,-5.9],
 ["ActivityNet","T-V",67.0,94.9,66.5,96.0,65.7,94.3,64.6,95.1,0.5],
 ["ActivityNet","T-VA",68.3,95.8,69.9,96.1,66.4,95.0,66.9,95.4,-1.6],
 ["VATEX","T-V",81.9,98.8,84.4,99.8,82.1,99.1,81.6,99.8,-2.5],
 ["VATEX","T-VA",85.9,99.3,87.0,99.5,84.7,99.1,84.6,100.0,-1.1],
 ["VATEX","T-VAS",86.8,99.3,87.7,100.0,84.9,99.1,84.2,99.8,-0.9],
]
build("Finetuning","Fine-tuning  ·  10 modes, own checkpoint  (R@1 / R@10)   [paper Table 2]",
 [("",2),("T2V Ours",2),("T2V Paper",2),("V2T Ours",2),("V2T Paper",2),("",1)],
 ["Benchmark","Mode","R@1","R@10","R@1","R@10","R@1","R@10","R@1","R@10","Δ T2V"],
 ft, ours_cols={3,4,7,8}, delta_cols={11}, merge_first=True)

out="/leonardo_work/IscrC_GMEG/anag0000/HyperAlign/HyperAlign_results.xlsx"
wb.save(out)
print(f"saved: {out}\nsheets: {wb.sheetnames}")
